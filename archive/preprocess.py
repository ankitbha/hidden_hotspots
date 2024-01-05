# ********************************************************************
# 
# Collection of functions to preprocessing tasks on the data.
#
# Date: Jan 28, 2020
#
# Author: Shiva R. Iyer
#
# ********************************************************************

import os
import argparse
import numpy as np
import pandas as pd

from tqdm import tqdm

def round_to_week(ts):
    """ Round a timestamp to the previous Monday. """
    #return (ts - pd.DateOffset(days=ts.dayofweek)).floor('D')
    return (ts - pd.DateOffset(days=ts.dayofweek, hour=0, minute=0, second=0))


def round_to_month(ts):
    """ Round a timestamp to beginning of the month. """
    #return (ts - pd.DateOffset(days=ts.day)).floor('D')
    return (ts - pd.DateOffset(days=ts.day-1, hour=0, minute=0, second=0))


# read data for an excel file and return a list of dataframes, 
# one for each monitor in the file
def read_xlsx(fpath):
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from datetime import datetime
    
    # handy conversion function to handle missing values
    def myconv(strval):
        if strval=='None' or strval==None:
            return np.nan
        else:
            return float(strval)

    workbook = load_workbook(fpath)
    sheet = workbook['Sheet1']
    fname = os.path.basename(fpath)
    
    #ROW_START = 14
    
    header = None
    for HEADER_ROW_NUM, row in enumerate(sheet.values, 1):
        if row[0] == 'From Date':
            header = [row[0]] + [s for s in row[2:] if s is not None]
            break
    
    # current row where we are
    rr = HEADER_ROW_NUM + 1
    
    # the sheet may have data from multiple locations, so the outer
    # loop iterates through the location, the inner loop iterates 
    # through the rows for each location
    #parts = fname[:fname.find('_' + suffix)].split('_')
    parts = fname[:-5].split('_')
    n_monitors = (len(parts) - 3) // 2
    print('Number of monitors:', n_monitors)
    
    data_list = []
    
    count = 0
    while count < n_monitors*2:
        name = '{}_{}'.format(parts[count], parts[count+1])
        
        data = []
        
        print('Reading data for', name)
        for row in sheet.iter_rows(rr, max_col=len(header)+1, values_only=True):
            date = row[0]
            if date is None or len(date) == 1:
                break
            readings = [myconv(val) for val in row[2:]]
            
            # to re-format the date and time in the same format as the Kaiterra dts
            dobj = datetime.strptime(date, '%d-%m-%Y %H:%M')
            dobj = pd.Timestamp(dobj, tz='Asia/Kolkata')
            data.append([dobj] + readings)
        
        df = pd.DataFrame(data, columns=header)
        df.rename(columns={'From Date':'timestamp_round', 'PM2.5':'pm25', 'PM10':'pm10'}, copy=False, inplace=True)
        df.set_index('timestamp_round', inplace=True)
        df.sort_index(axis=1, inplace=True) # sort the columns
        
        data_list.append((name, df))
        
        rr += len(data) + 2
        count += 2
    
    return data_list


def average_govdata(fpath, interval, start_dt=None, end_dt=None):

    if not os.path.exists(fpath):
        raise FileNotFoundError('Input file \"{}\" not found!'.format(fpath))
    
    # read in the data with index as the tuple (id, timestamp)
    data = pd.read_csv(fpath, index_col=[0,1], parse_dates=True)

    # pm25 has negative values, so replace them with NaNs
    data.pm25.loc[data.pm25 < 0] = np.nan

    # sort the df by ids first, time next
    data.sort_index(level=0, inplace=True)
    n_mid = data.index.levels[0].size
    
    # set the timezone in data to be UTC so that the df becomes tz-aware
    if data.index.levels[1].tz is None:
        data = data.tz_localize('UTC', level=1, copy=False)
    
    # now, the shift the timestamps in the data to the new timezone
    # (can be done only in tz-aware df)
    data = data.tz_convert('Asia/Kolkata', level=1, copy=False)
    
    # start and end dates for data
    if start_dt is None:
        start_dt = data.index.levels[1][0]

    if end_dt is None:
        end_dt = data.index.levels[1][-1]
    
    # if start & end dates are not in IST, then convert them to IST,
    # assuming they are in UTC if they are not tz-aware
    if start_dt.tzname != 'IST':
        if start_dt.tzinfo is None:
            start_dt = start_dt.tz_localize('UTC')
        start_dt = start_dt.tz_convert('Asia/Kolkata')
    
    if end_dt.tzname != 'IST':
        if end_dt.tzinfo is None: 
            end_dt = end_dt.tz_localize('UTC')
        end_dt = end_dt.tz_convert('Asia/Kolkata')

    start_dt = start_dt.floor('D')
    end_dt = end_dt.ceil('D')

    # now, filter through the start and end dates
    data = data.loc[(slice(None), slice(start_dt, end_dt)),:]

    # for the averaging operation, first append a column with rounded timestamps
    data.reset_index(inplace=True)
    data.rename({'timestamp_round' : 'timestamp'}, axis=1, inplace=True)

    # reindexing so that all timestamps are in same range for every sensor
    if 'W' in interval:
        interval = '1W'
        data['timestamp_round'] = data.timestamp.apply(round_to_week, convert_dtype=False)
        start = round_to_week(start_dt)
        end = round_to_week(end_dt)
        if (end_dt - end).value // 1e9 > 0:
            end += pd.DateOffset(weeks=1)
        newindex = pd.date_range(start, end, freq=pd.offsets.Week(weekday=0), closed='left', tz='Asia/Kolkata', name='timestamp_round')
    elif 'M' in interval:
        interval = '1M'
        data['timestamp_round'] = data.timestamp.apply(round_to_month, convert_dtype=False)
        start = round_to_month(start_dt)
        end = round_to_month(end_dt)
        if (end_dt - end).value // 1e9 > 0:
            end += pd.DateOffset(months=1)
        newindex = pd.date_range(start, end, freq=pd.offsets.MonthBegin(), closed='left', tz='Asia/Kolkata', name='timestamp_round')
    else:
        data['timestamp_round'] = data.timestamp.apply(lambda ts: ts.floor(interval), convert_dtype=False)
        start, end = start_dt, end_dt
        newindex = pd.date_range(start, end, freq=interval, closed='left', tz='Asia/Kolkata', name='timestamp_round')

    # create save paths and directory (if required)
    saveprefix = 'govdata_{}_{}_{}'.format(interval, start_dt.strftime('%Y%m%d'), end_dt.strftime('%Y%m%d'))
    savedir = os.path.dirname(fpath)
    subdir = os.path.join(savedir, saveprefix)
    if not os.path.exists(subdir):
        os.mkdir(subdir)

    # do the averaging for each monitor id separately and save it
    grouped = data.groupby('monitor_id')
    groups = []
    for name, group in tqdm(grouped, total=n_mid, desc='Avging for each monitor'):
        df = group.drop(['monitor_id', 'timestamp'], axis=1).groupby('timestamp_round').mean()
        df = df.reindex(newindex, axis=0)
        df.to_csv(os.path.join(subdir, name + '.csv'))
        
        df['monitor_id'] = name
        df.reset_index(inplace=True)
        df.set_index(['monitor_id', 'timestamp_round'], inplace=True)
        groups.append(df)

    data_avged = pd.concat(groups)

    # save the final big concatenated dataframe
    data_avged.to_csv(os.path.join(savedir, saveprefix + '.csv'), float_format='%.2f')

    return data_avged


def average_kaiterra_data(fpath, interval, start_dt=None, end_dt=None):

    if not os.path.exists(fpath):
        raise FileNotFoundError('Input file \"{}\" not found!'.format(fpath))
    
    # read in the data with index as the tuple (id, timestamp)
    data = pd.read_csv(fpath, index_col=[2,0], parse_dates=True)
    
    # sort the df by ids first, time next
    data.sort_index(level=0, inplace=True)
    n_mid = data.index.levels[0].size
    
    # set the timezone in data to be UTC so that the df becomes tz-aware
    if data.index.levels[1].tz is None:
        data = data.tz_localize('UTC', level=1, copy=False)
    
    # now, the shift the timestamps in the data to the new timezone
    # (can be done only in tz-aware df)
    data = data.tz_convert('Asia/Kolkata', level=1, copy=False)
    
    # start and end dates for data
    if start_dt is None:
        start_dt = data.index.levels[1][0]

    if end_dt is None:
        end_dt = data.index.levels[1][-1]
    
    # if start & end dates are not in IST, then convert them to IST,
    # assuming they are in UTC
    if start_dt.tzname != 'IST':
        if start_dt.tzinfo is None:
            start_dt = start_dt.tz_localize('UTC')
        start_dt = start_dt.tz_convert('Asia/Kolkata')
    
    if end_dt.tzname != 'IST':
        if end_dt.tzinfo is None: 
            end_dt = end_dt.tz_localize('UTC')
        end_dt = end_dt.tz_convert('Asia/Kolkata')

    start_dt = start_dt.floor('D')
    end_dt = end_dt.ceil('D')
    
    # rename the short ids to make them uppercase
    # --> this op takes time
    data.rename(str.upper, axis=0, level=0, inplace=True)

    # now, filter through the start and end dates
    data = data.loc[(slice(None), slice(start_dt, end_dt)),:]

    # for the averaging operation, first append a column with rounded timestamps
    data.reset_index(inplace=True)

    # reindexing so that all timestamps are in same range for every sensor
    if 'W' in interval:
        interval = '1W'
        start = round_to_week(start_dt)
        end = round_to_week(end_dt)
        if (end_dt - end).value // 1e9 > 0:
            end += pd.DateOffset(weeks=1)
        newindex = pd.date_range(start, end, freq=pd.offsets.Week(weekday=0), closed='left', tz='Asia/Kolkata', name='timestamp_round')
        data['timestamp_round'] = data.time.apply(round_to_week, convert_dtype=False)
    elif 'M' in interval:
        interval = '1M'
        start = round_to_month(start_dt)
        end = round_to_month(end_dt)
        if (end_dt - end).value // 1e9 > 0:
            end += pd.DateOffset(months=1)
        newindex = pd.date_range(start, end, freq=pd.offsets.MonthBegin(), closed='left', tz='Asia/Kolkata', name='timestamp_round')
        data['timestamp_round'] = data.time.apply(round_to_month, convert_dtype=False)
    else:
        data['timestamp_round'] = data.time.apply(lambda ts: ts.floor(interval), convert_dtype=False)
        start, end = start_dt, end_dt
        newindex = pd.date_range(start, end, freq=interval, closed='left', tz='Asia/Kolkata', name='timestamp_round')

    # do the averaging for each monitor id separately
    grouped = data.groupby('short_id')
    groups = []
    for name, group in tqdm(grouped, total=n_mid, desc='Avging for each monitor'):
        df = group.groupby('timestamp_round')[['pm_25', 'pm_10']].mean()
        df = df.reindex(newindex, axis=0)
        df['field_egg_id'] = name
        df['device_udid'] = group.device_udid.iloc[0]
        df.reset_index(inplace=True)
        df.set_index(['field_egg_id', 'timestamp_round'], inplace=True)
        groups.append(df)

    data_avged = pd.concat(groups)
    data_avged.rename({'pm_25':'pm25', 'pm_10':'pm10'}, axis=1, inplace=True)

    data_final = data_avged[['pm25', 'pm10']]

    # save the final dataframe
    saveprefix = 'kaiterra_fieldeggid_{}_{}_{}'.format(interval, start_dt.strftime('%Y%m%d'), end_dt.strftime('%Y%m%d'))
    savedir = os.path.dirname(fpath)
    data_final.to_csv(os.path.join(savedir, saveprefix + '.csv'))

    # also save the data for each sensor separately
    subdir = os.path.join(savedir, saveprefix)
    if not os.path.exists(subdir):
        os.mkdir(subdir)

    grouped = data_final.groupby(level=0)
    for name, group in grouped:
        group.reset_index(level=0, drop=True, inplace=True)
        group.to_csv(os.path.join(subdir, name + '.csv'))

    return data_avged


if __name__ == '__main__':

    ts_type = lambda arg: pd.Timestamp(arg, tz='Asia/Kolkata')

    parser = argparse.ArgumentParser()

    subparsers = parser.add_subparsers(title='Preprocessing operations on raw data',
                                       help='List of valid preprocessing operations',
                                       dest='operation')
    parser_a = subparsers.add_parser('average', aliases=['avg'])
    parser_a.add_argument('source', choices=('kaiterra', 'govdata'), help='Source (kaiterra or govdata)')
    parser_a.add_argument('interval', help='Averaging interval (e.g. 15T, 1H, 1D)')
    parser_a.add_argument('--file', '-f', help='Full path to data file')
    parser_a.add_argument('--start', type=ts_type, help='Start timestamp in IST (=UTC+5:30)')
    parser_a.add_argument('--end', type=ts_type, help='End timestamp in IST (=UTC+5:30)')

    args = parser.parse_args()

    if args.operation is None:
        print('At least one preprocessing operation should be specified.')
        parser.print_usage()

    elif args.operation in ('average', 'avg'):
        if args.source == 'kaiterra':
            if args.file is None:
                fpath_default = os.path.join('data', 'kaiterra', 'kaiterra_fieldeggid_all_current.csv')
                print('Input file not provided. Using the default: "{}"'.format(fpath_default))
                args.file = fpath_default
            data_avged = average_kaiterra_data(args.file, args.interval, args.start, args.end)
        else:
            if args.file is None:
                fpath_default = os.path.join('data', 'govdata', 'govdata_current.csv')
                print('Input file not provided. Using the default: "{}"'.format(fpath_default))
                args.file = fpath_default
            data_avged = average_govdata(args.file, args.interval, args.start, args.end)